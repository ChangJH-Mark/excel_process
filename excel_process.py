import openpyxl as xl
CHKS_INDEX='I'
workbook = xl.Workbook()
ws = workbook.active
chks = list()
task = "SINGLE"
file = "C:/Users/mark/Desktop/Datas/无预读分析.xlsx"

def chks_valid(col):
    chks = list()
    for cell in col:
        if cell.value not in chks:
            chks.append(cell.value)
    chks.pop(0)
    for i in range(min(chks),max(chks)+1):
        if i not in chks:
            print("CHKS ERROR")
            break
    return chks

def single_cost(chks):
    print("Process Single Cost")
    start,end = 0.0,0.0
    dts = dict()
    for i in range(2,ws.max_row+1):
        chk = ws.cell(row=i,column=9).value
        time = ws.cell(row=i,column=5).value
        status = ws.cell(row=i,column=16).value
        if status == "Create_Thread" or status =="SUCCESS_SIZE=131072":
            if status == "Create_Thread":
                if chk in dts:
                    print("len 32")
                    assert(False)
                dts[chk]=[time]
            if status == "SUCCESS_SIZE=131072":
                if chk not in dts or len(dts[chk])>1:
                    print("len 37")
                    assert(False)
                dts[chk].extend([time,i])
        else:
            continue
    for k in dts.keys():
        if len(dts[k])!=3:
            print("error:" + " \t".join(dts[k]))
        l = dts[k]
        ws.cell(row=l[2],column=1,value=(l[1] - l[0]))

def main():
    global workbook,ws,chks,task,file
    workbook = xl.load_workbook(file)
    ws = workbook.active
    chks = chks_valid(ws[CHKS_INDEX])
    if "SINGLE" in task:
        single_cost(chks)
    workbook.save(file)
    print("All Processed Well")

if __name__ == "__main__":
    main()