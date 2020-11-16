import openpyxl as xl
map = {"singlecost":1,"time":2,"TID":3,"function":4,"chk:":5,"CHUNCK":6,"size:":7,
       "SIZE":8,"off:":9,"OFF":10,"file:":11,"FILE":12,"STATUS":13}
def detect_clear_useful_chk(ws):
    start_chk = 0
    for row in range(2,ws.max_row+1):
        if ws.cell(row=row, column=map["function"]).value=="create_chuncks":
            if start_chk == 0:
                start_chk = ws.cell(row=row,column=map["CHUNCK"]).value
            else:
                if start_chk != ws.cell(row=row,column=map["CHUNCK"]).value:
                    print(row)
                    assert(False)
                start_chk = 0
        elif ws.cell(row=row, column=map["function"]).value=="clear_chunck":
            chk = int(ws.cell(row=row,column=map["CHUNCK"]).value)
            if abs(chk-int(start_chk)) < 70:
                print(row)
                assert(False)
    print("yes of course!")

def cal_reqs_interns(ws):
    start_time = 0
    end_time = 0
    count=0
    res = {}
    for row in range(2,ws.max_row+1):
        chk = int(ws.cell(row=row,column=map["CHUNCK"]).value)
        status = ws.cell(row=row,column=map["STATUS"]).value
        time = ws.cell(row=row,column=map["time"]).value
        if status == "OVER_FETCHED=131072":
            count = count + 1
            start_time = time
        if status == "Create_Thread" and count == 8:
            end_time = time
            res[row]=end_time - start_time
            count = 0
    return res

def cal_reqs_costs(ws):
    start_time = 0
    end_time = 0
    count = 0
    res = {}
    for row in range(2,ws.max_row+1):
        chk = int(ws.cell(row=row,column=map["CHUNCK"]).value)
        status = ws.cell(row=row,column=map["STATUS"]).value
        time = ws.cell(row=row,column=map["time"]).value
        if count == 0 and status == "Create_Thread":
            start_time = time
        elif status == "SUCCESS_SIZE=131072":
            count = count + 1
            if count == 8:
                end_time = time
                res[row]=end_time-start_time
                count = 0
                start_time = 0
    return res

def cal_clear_chk(ws):
    start_time = 0
    end_time = 0
    res={}
    for row in range(2,ws.max_row+1):
        if ws.cell(row=row, column=4).value == "create_chuncks":
            if start_time == 0:
                start_time = ws.cell(row=row,column=2).value
            else:
                end_time=ws.cell(row=row,column=2).value
                res[row]=(end_time-start_time)
                start_time=0
                end_time=0
    return res

def chks_value_record(ws,key_col=6,status=[]):
    '''
    :param ws: workbook sheet
    :param key_col: col value used as dict key
    :param status: start status for start record and end status for end record
    :return:
    '''
    res = dict()
    for i in range(2,ws.max_row+1):
        time = ws.cell(row=i, column=2).value
        key = ws.cell(row=i,column=key_col).value
        st = ws.cell(row=i,column=13).value
        if key not in res:
            if st == status[0]:
                res[key]=[time,i]
        elif st == status[1]:
            if(len(res[key])!=2):
                print(key, res[key], i)
                assert(False)
            res[key].extend([time,i])

        else:
            pass
    for key in res.keys():
        if len(res[key]) != 4:
            print(key,res[key])
            assert(False)
    return res


def main():
    file="C:/Users/mark/Desktop/RA_seq_2M.xlsx"
    #file = "C:/Users/mark/Desktop/NRA_seq_2M.xlsx"
    work = xl.load_workbook(file)
    ws = work.active
    '''检测clear是否有问题'''
    detect_clear_useful_chk(ws)
    '''记录create_chks消耗的时间'''
    '''记录reqs interns消耗的时间'''
    '''
    res=cal_reqs_costs(ws)
    for k in res.keys():
        v = res[k]
        ws.cell(row=k,column=map["STATUS"]+1,value=v)
    work.save(file)
    '''
'''
    infos = chks_value_record(ws,10,["Create_Thread","SUCCESS_SIZE=8192"])
    for off in infos.keys():
        start = infos[off][0]
        start_row = infos[off][1]
        end_time = infos[off][2]
        end_row = infos[off][3]
        if str(int(off)-8192) in infos:
            lastone = infos[str(int(off)-8192)]
            ws.cell(row=start_row,column=1,value=start-lastone[2])'''

if __name__ == "__main__":
    main()