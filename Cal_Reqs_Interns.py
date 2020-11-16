import openpyxl as xl
map = {"singlecost":1,"time":2,"TID":3,"function":4,"chk:":5,"CHUNCK":6,"size:":7,
       "SIZE":8,"off:":9,"OFF":10,"file:":11,"FILE":12,"STATUS":13}

def cal_reqs_interns(ws):
    start_time = 0
    end_time = 0
    count=0
    res = {}
    for row in range(2,ws.max_row+1):
        chk = int(ws.cell(row=row,column=map["CHUNCK"]).value)
        status = ws.cell(row=row,column=map["STATUS"]).value
        time = ws.cell(row=row,column=map["time"]).value
        if status == "OVER_FETCHED=131072":
            count = count + 1
            start_time = time
        if status == "Create_Thread" and count == 8:
            end_time = time
            res[row]=end_time - start_time
            count = 0
    return res
def main():
    file = "C:/Users/mark/Desktop/RA_seq_2M.xlsx"
    work = xl.load_workbook(file)
    ws = work.active
    res = cal_reqs_interns(ws)
    for k in res.keys():
        v = res[k]
        ws.cell(row=k,column=map["singlecost"],value=v)
    work.save(file)
if __name__ == "__main__":
    main()