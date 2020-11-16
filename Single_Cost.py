import openpyxl as xl
map = {"singlecost":1,"time":2,"TID":3,"function":4,"chk:":5,"CHUNCK":6,"size:":7,
       "SIZE":8,"off:":9,"OFF":10,"file:":11,"FILE":12,"STATUS":13}
def cal_single_cost(ws):
    start_times={}
    res = {}
    for row in range(2,ws.max_row+1):
        chk = int(ws.cell(row=row,column=map["CHUNCK"]).value)
        status = ws.cell(row=row,column=map["STATUS"]).value
        time = ws.cell(row=row,column=map["time"]).value
        if status == "SEND":
            if chk in start_times:
                print(row,"wrong start_time")
            start_times[chk] = time
        if status == "ARRIVED":
            if chk in start_times:
                res[row]=time-start_times[chk]
            else:
                print(row)
    return res

def main():
    print("fuckyou")
    file = "C:/Users/mark/Desktop/RA_seq_2M.xlsx"
    work = xl.load_workbook(file)
    ws = work.active
    ws2 = work['Send_Arrived']
    print("File: " + file)
    print("sheet:" + ws2.title)
    c = input("OK?")
    if c == "N" or c == "n":
        assert(False)
    res = cal_single_cost(ws2)
    for k in res.keys():
        v = res[k]
        ws2.cell(row=k,column=map["singlecost"],value=v)
    work.save(file)
if __name__ == "__main__":
    main()