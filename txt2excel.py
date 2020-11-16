import re
import openpyxl as xl
import fileinput
def takeElement(e):
    return e[0]
headers = ["singlecost","time","TID","function","chk:","CHUNCK","size:","SIZE","off:","OFF","file:","FILE","STATUS"]
start_hour = 20
start_min = 50
start_sec = 50
workbook = xl.Workbook()
ws = workbook.active
file = "C:/Users/mark/Desktop/RA.xlsx"
infile ="C:/Users/mark/Desktop/ra_analyze.log"

for i in range(0, len(headers)):
    ws.cell(row=1, column=i + 1, value=headers[i])
threads = {}
num=1
with open(infile,"r") as f:
    finals=[]
    while True:
        line = f.readline()
        if not line:
            break
        print(num)
        num = num +1
        line = line.split()
        nline=[]
        if not (len(line)>5 and line[5]=="chk:"):
            continue
        t=re.match('[0-9]+:[0-9]+:[0-9]+\.[0-9]+',line[1]).group(0)
        t=re.split('\:|\.',t)
        hour = int(t[0])
        min = int(t[1])
        sec = int(hour - start_hour)*60*60 + int(min - start_min) * 60 + int(t[2]) - start_sec
        millsec=int(t[3])
        millsecs=(sec * 1000 + millsec / 1000)
        nline.append(millsecs)
        if re.match('[0-9abcdef]+',line[2]):
            if line[2] not in threads:
                threads[line[2]] = str(1+len(threads))
            nline.append(threads[line[2]])
        for i in line[4:13]:
            nline.append(i)
        nline[-1]=nline[-1].split('/')[-1]
        nline.append("_".join(line[13:]))
        finals.append(nline)
    finals.sort(key=takeElement)
    num = 1
    for r in range(0,len(finals)):
        row = finals[r]
        print(num)
        num = num + 1
        for col in range(len(headers) - len(row),len(headers)):
            ws.cell(row = r+2,column=col + 1, value=row[col + len(row)-len(headers)])
    workbook.save(file)